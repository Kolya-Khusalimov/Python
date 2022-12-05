import openpyxl
import datetime

wb = openpyxl.load_workbook("23.13.xlsx")
ids = wb["Авіакомпанії"]
airp = wb["Аеропорти"]
flights = wb["Рейси"]
flight_to = "Париж"
id_dict = {}
airp_dict = {}

days_dict = {'1': 'Monday',
             '2': 'Tuesday',
             '3': 'Wednesday',
             '4': 'Thursday',
             '5': 'Friday',
             '6': 'Saturday',
             '7': 'Sunday'
             }

cost_dict = {}
time_dict = {}
min_col = ids.min_column

for i in range(ids.min_row + 1, ids.max_row + 1):
    c1 = ids.cell(i, min_col)
    c2 = ids.cell(i, min_col + 1)
    id_dict[c1.value] = c2.value



min_col = airp.min_column
for i in range(airp.min_row + 1, airp.max_row + 1):
    c1 = airp.cell(i, min_col)
    c2 = airp.cell(i, min_col + 1)
    c3 = airp.cell(i, min_col + 2)
    airp_dict[c1.value] = [c2.value, c3.value]

for key, value in airp_dict.items():
    if value[1] == "Париж":
        flight_to = key

min_col = flights.min_column
for i in range(flights.min_row + 1, flights.max_row + 1):
    from_id = flights.cell(i, min_col).value
    to_id = flights.cell(i, min_col + 1).value
    flight_id = flights.cell(i, min_col + 2).value
    days = str(flights.cell(i, min_col + 3).value)
    depart = flights.cell(i, min_col + 4).value
    arrive = flights.cell(i, min_col + 5).value
    flight_class = flights.cell(i, min_col + 6).value
    cost = flights.cell(i, min_col + 7).value

    cost_dict[(from_id, to_id, flight_id, flight_class)] = cost

    '''new_days=[]
    days=list(days)
    for day in days:
        if day in days_dict:
            new_days.append(day)'''

    td = depart.hour * 60 + depart.minute
    ta = arrive.hour * 60 + arrive.minute
    delta = abs(ta - td)
    time_dict[(from_id, to_id, flight_id, flight_class, str(depart), str(arrive))] = delta

min_cost = max(cost_dict.values())
min_cost_dict = {}
for key, value in cost_dict.items():
    if key[1] == flight_to:
        if value <= min_cost:
            min_cost_dict.clear()
            min_cost_dict[key] = value

print(min_cost_dict)
min_time = max(time_dict.values())
min_time_dict = {}
for key, value in time_dict.items():
    if key[1] == flight_to:
        if value <= min_time:
            min_time_dict.clear()
            min_time_dict[key] = value
print(min_time_dict)